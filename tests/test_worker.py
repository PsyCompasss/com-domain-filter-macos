import tempfile
import threading
import unittest
from pathlib import Path

from openpyxl import load_workbook

from com_domain_filter.cloudflare import (
    QueryResult,
    STATUS_EXACT_AVAILABLE,
    TransientPageError,
    VerificationRequired,
)
from com_domain_filter.excel_store import SHEET_NAME
from com_domain_filter.storage import HistoryStore
from com_domain_filter.worker import RunConfig, SearchWorker


class FakeChecker:
    instances = []

    def __init__(self, site_url, profile_dir):
        self.site_url = site_url
        self.keep_browser = None
        type(self).instances.append(self)

    def start(self):
        pass

    def close(self, keep_browser=False):
        self.keep_browser = keep_browser

    def verification_present(self):
        return False

    def query(self, domain):
        return QueryResult(domain, STATUS_EXACT_AVAILABLE, domain, True, True)

    def recover_page(self):
        pass


class VerificationLoopChecker(FakeChecker):
    def query(self, domain):
        raise VerificationRequired("需要验证")

    def wait_for_verification(self, stop_event):
        return True


class BlockingChecker(FakeChecker):
    started = threading.Event()
    release = threading.Event()

    def query(self, domain):
        type(self).started.set()
        type(self).release.wait(timeout=5)
        return super().query(domain)


class TransientThenSuccessChecker(FakeChecker):
    query_calls = 0
    recover_calls = 0

    def query(self, domain):
        type(self).query_calls += 1
        if type(self).query_calls == 1:
            raise TransientPageError("页面暂时空白")
        return super().query(domain)

    def recover_page(self):
        type(self).recover_calls += 1


class AlwaysTransientChecker(FakeChecker):
    query_calls = 0

    def query(self, domain):
        type(self).query_calls += 1
        raise TransientPageError("持续空白")


class StartupTransientChecker(FakeChecker):
    start_calls = 0

    def start(self):
        type(self).start_calls += 1
        if type(self).start_calls == 1:
            raise TransientPageError("网站暂时打不开")


class FastWaitEvent(threading.Event):
    def wait(self, timeout=None):
        return super().wait(0)


class WorkerTests(unittest.TestCase):
    @staticmethod
    def make_config(root: Path, limit_tests: int = 2) -> RunConfig:
        return RunConfig(
            site_url="https://domains.cloudflare.com/",
            characters=("a", "b", "c"),
            patterns=("AAA",),
            prefix="abc",
            suffix="",
            unlimited_length=6,
            interval_seconds=0.01,
            retry_interval_seconds=0.01,
            limit_tests_enabled=True,
            limit_tests=limit_tests,
            limit_found_enabled=False,
            limit_found=99,
            run_until_stopped=False,
            excel_path=root / "results.xlsx",
            profile_dir=root / "profile",
        )

    def test_worker_stops_at_test_limit_and_writes_excel(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            events = []
            config = self.make_config(root)
            worker = SearchWorker(
                config,
                HistoryStore(root / "state.db"),
                lambda kind, payload: events.append((kind, payload)),
                checker_factory=FakeChecker,
            )
            worker.start()
            worker.thread.join(timeout=5)
            self.assertFalse(worker.is_alive)
            self.assertEqual(worker.checked, 2)
            self.assertEqual(worker.found, 2)
            self.assertTrue(any(kind == "finished" for kind, _ in events))
            self.assertTrue(FakeChecker.instances[-1].keep_browser)
            workbook = load_workbook(config.excel_path)
            self.assertEqual(workbook[SHEET_NAME].max_row, 3)

    def test_manual_stop_keeps_browser_open(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            events = []
            BlockingChecker.started.clear()
            BlockingChecker.release.clear()
            history = HistoryStore(root / "state.db")
            worker = SearchWorker(
                self.make_config(root, limit_tests=10),
                history,
                lambda kind, payload: events.append((kind, payload)),
                checker_factory=BlockingChecker,
            )
            worker.start()
            self.assertTrue(BlockingChecker.started.wait(timeout=5))
            worker.stop(keep_browser_open=True)
            BlockingChecker.release.set()
            worker.thread.join(timeout=5)

            self.assertFalse(worker.is_alive)
            self.assertTrue(BlockingChecker.instances[-1].keep_browser)
            self.assertEqual(worker.checked, 1)
            self.assertEqual(history.total_count(), 1)
            messages = [payload["message"] for kind, payload in events if kind == "finished"]
            self.assertIn("已手动停止；Chrome浏览器保持打开", messages)

    def test_domain_is_reserved_while_query_is_still_running(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            events = []
            BlockingChecker.started.clear()
            BlockingChecker.release.clear()
            history = HistoryStore(root / "state.db")
            worker = SearchWorker(
                self.make_config(root, limit_tests=10),
                history,
                lambda kind, payload: events.append((kind, payload)),
                checker_factory=BlockingChecker,
            )
            worker.start()
            self.assertTrue(BlockingChecker.started.wait(timeout=5))
            current = next(payload["domain"] for kind, payload in events if kind == "current")
            self.assertTrue(history.has_tested(current))
            worker.stop(keep_browser_open=True)
            BlockingChecker.release.set()
            worker.thread.join(timeout=5)

    def test_worker_stops_after_repeated_verification(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            events = []
            worker = SearchWorker(
                self.make_config(root, limit_tests=1),
                HistoryStore(root / "state.db"),
                lambda kind, payload: events.append((kind, payload)),
                checker_factory=VerificationLoopChecker,
            )
            worker.start()
            worker.thread.join(timeout=5)
            self.assertFalse(worker.is_alive)
            errors = [payload["message"] for kind, payload in events if kind == "error"]
            self.assertEqual(len(errors), 1)
            self.assertIn("连续要求真人验证", errors[0])

    def test_transient_query_failure_refreshes_without_error_popup(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            events = []
            TransientThenSuccessChecker.query_calls = 0
            TransientThenSuccessChecker.recover_calls = 0
            worker = SearchWorker(
                self.make_config(root, limit_tests=1),
                HistoryStore(root / "state.db"),
                lambda kind, payload: events.append((kind, payload)),
                checker_factory=TransientThenSuccessChecker,
            )
            worker.stop_event = FastWaitEvent()
            worker.start()
            worker.thread.join(timeout=5)

            self.assertFalse(worker.is_alive)
            self.assertEqual(TransientThenSuccessChecker.query_calls, 2)
            self.assertEqual(TransientThenSuccessChecker.recover_calls, 1)
            self.assertFalse(any(kind in {"error", "verification"} for kind, _ in events))
            self.assertTrue(
                any(kind == "status" and "自动刷新继续" in payload["message"] for kind, payload in events)
            )

    def test_single_domain_is_skipped_after_three_transient_failures(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            events = []
            AlwaysTransientChecker.query_calls = 0
            history = HistoryStore(root / "state.db")
            worker = SearchWorker(
                self.make_config(root, limit_tests=1),
                history,
                lambda kind, payload: events.append((kind, payload)),
                checker_factory=AlwaysTransientChecker,
            )
            worker.stop_event = FastWaitEvent()
            worker.start()
            worker.thread.join(timeout=5)

            self.assertFalse(worker.is_alive)
            self.assertEqual(AlwaysTransientChecker.query_calls, 3)
            self.assertEqual(worker.checked, 1)
            self.assertTrue(any(kind == "status" and "已跳过" in payload["message"] for kind, payload in events))

    def test_transient_startup_failure_retries_without_error_popup(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            events = []
            StartupTransientChecker.start_calls = 0
            worker = SearchWorker(
                self.make_config(root, limit_tests=1),
                HistoryStore(root / "state.db"),
                lambda kind, payload: events.append((kind, payload)),
                checker_factory=StartupTransientChecker,
            )
            worker.stop_event = FastWaitEvent()
            worker.start()
            worker.thread.join(timeout=5)

            self.assertFalse(worker.is_alive)
            self.assertEqual(StartupTransientChecker.start_calls, 2)
            self.assertFalse(any(kind in {"error", "verification"} for kind, _ in events))
            self.assertTrue(
                any(kind == "status" and "Chrome连接暂时中断" in payload["message"] for kind, payload in events)
            )


if __name__ == "__main__":
    unittest.main()
