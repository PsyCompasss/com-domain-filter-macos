import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from com_domain_filter.excel_store import (
    HISTORY_HEADERS,
    HISTORY_SHEET_NAME,
    ExcelStore,
    HEADERS,
    HistoryExcelStore,
    SHEET_NAME,
)
from com_domain_filter.storage import HistoryStore, SettingsStore


class StorageAndExcelTests(unittest.TestCase):
    def test_history_deduplicates(self):
        with tempfile.TemporaryDirectory() as temp:
            store = HistoryStore(Path(temp) / "state.db")
            args = ("abc.com", "exact_unavailable", "2026-08-21T12:00:00+08:00", "AAA", "", "", "")
            self.assertTrue(store.record(*args))
            self.assertFalse(store.record(*args))
            self.assertTrue(store.has_tested("ABC.COM"))
            self.assertEqual(store.total_count(), 1)

    def test_idn_history_treats_readable_and_punycode_domains_as_the_same_record(self):
        with tempfile.TemporaryDirectory() as temp:
            store = HistoryStore(Path(temp) / "state.db")
            args = ("abc.中国", "exact_available", "2026-08-21T12:00:00+08:00", "不限", "", "", "")
            self.assertTrue(store.record(*args))
            self.assertFalse(store.record("abc.xn--fiqs8s", *args[1:]))
            self.assertTrue(store.has_tested("abc.xn--fiqs8s"))
            self.assertEqual(store.history_rows()[0][0], "abc.中国")
            self.assertEqual(store.delete_domains(["abc.xn--fiqs8s"]), 1)

    def test_excel_saves_readable_idn_and_deduplicates_punycode(self):
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "results.xlsx"
            store = ExcelStore(path)
            rest = ("2026-08-21T12:00:00+08:00", "不限", "", "", "万网")
            self.assertTrue(store.append_if_new("abc.xn--fiqs8s", *rest))
            self.assertFalse(store.append_if_new("abc.中国", *rest))
            sheet = load_workbook(path)[SHEET_NAME]
            self.assertEqual(sheet["A2"].value, "abc.中国")

    def test_query_is_reserved_before_result_and_then_finalized(self):
        with tempfile.TemporaryDirectory() as temp:
            store = HistoryStore(Path(temp) / "state.db")
            self.assertTrue(store.reserve("abc.com", "start", "AAA", "", ""))
            self.assertTrue(store.has_tested("abc.com"))
            self.assertFalse(store.reserve("abc.com", "start", "AAA", "", ""))
            self.assertTrue(
                store.finalize("abc.com", "exact_unavailable", "finish", "AAA", "", "", "done")
            )
            with store._connect() as connection:
                row = connection.execute(
                    "SELECT status, checked_at, detail FROM tested_domains WHERE domain = 'abc.com'"
                ).fetchone()
            self.assertEqual(row, ("exact_unavailable", "finish", "done"))

    def test_settings_round_trip(self):
        with tempfile.TemporaryDirectory() as temp:
            store = SettingsStore(Path(temp) / "settings.json")
            store.save({"patterns": ["AAA"], "interval": "5"})
            self.assertEqual(store.load()["patterns"], ["AAA"])

    def test_history_can_be_listed_deleted_and_cleared(self):
        with tempfile.TemporaryDirectory() as temp:
            store = HistoryStore(Path(temp) / "state.db")
            store.record("one.com", "exact_unavailable", "2026-08-21T12:00:00+08:00", "AAA", "", "")
            store.record("two.com", "exact_available", "2026-08-21T13:00:00+08:00", "ABC", "", "")
            self.assertEqual([row[0] for row in store.history_rows()], ["two.com", "one.com"])
            self.assertEqual(store.delete_domains(["ONE.COM"]), 1)
            self.assertFalse(store.has_tested("one.com"))
            self.assertEqual(store.clear(), 1)
            self.assertEqual(store.total_count(), 0)

    def test_excel_append_and_deduplicate(self):
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "results.xlsx"
            store = ExcelStore(path)
            args = (
                "abc123.com",
                "2026-08-21T12:00:00+08:00",
                "ABC",
                "",
                "",
                "https://domains.cloudflare.com/",
            )
            self.assertTrue(store.append_if_new(*args))
            self.assertFalse(store.append_if_new(*args))
            workbook = load_workbook(path)
            sheet = workbook[SHEET_NAME]
            self.assertEqual(tuple(cell.value for cell in sheet[1]), HEADERS)
            self.assertEqual(sheet.max_row, 2)
            self.assertEqual(sheet["A2"].value, "abc123.com")
            self.assertEqual(sheet.freeze_panes, "A2")

    def test_history_excel_exports_all_statuses(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            history = HistoryStore(root / "state.db")
            history.record(
                "abc.com",
                "exact_available",
                "2026-08-21T12:00:00+08:00",
                "AAA",
                "",
                "",
                "confirmed",
                "https://wanwang.aliyun.com/domain",
            )
            history.record(
                "xyz.com",
                "exact_unavailable",
                "2026-08-21T12:01:00+08:00",
                "ABC",
                "",
                "",
                "registered",
                "https://wanwang.aliyun.com/domain",
            )
            path = root / "history.xlsx"
            self.assertEqual(HistoryExcelStore(path).export(history.history_rows()), 2)
            workbook = load_workbook(path)
            sheet = workbook[HISTORY_SHEET_NAME]
            self.assertEqual(tuple(cell.value for cell in sheet[1]), HISTORY_HEADERS)
            self.assertEqual(sheet.max_row, 3)
            self.assertEqual(sheet.freeze_panes, "A2")
            self.assertEqual(sheet.auto_filter.ref, "A1:H3")


if __name__ == "__main__":
    unittest.main()
