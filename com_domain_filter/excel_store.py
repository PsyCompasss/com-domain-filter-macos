from __future__ import annotations

import os
import tempfile
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


SHEET_NAME = "可注册COM域名"
HEADERS = ("域名", "查询时间", "规律", "固定开头", "固定结尾", "查询网站")
HISTORY_SHEET_NAME = "已查询记录"
HISTORY_HEADERS = ("域名", "查询状态", "查询时间", "规律组合", "固定开头", "固定结尾", "查询网站", "详细信息")
STATUS_LABELS = {
    "exact_available": "名称完全一致且可注册",
    "available_mismatch": "可注册但名称不一致",
    "exact_unavailable": "名称完全一致但已注册",
    "no_com": "未确认到 .com 结果",
    "query_started": "查询未完成",
    "query_failed": "查询失败",
}


class ExcelStoreError(RuntimeError):
    pass


class ExcelStore:
    def __init__(self, path: Path) -> None:
        self.path = Path(path).expanduser()
        if self.path.suffix.lower() != ".xlsx":
            raise ExcelStoreError("结果文件必须使用 .xlsx 后缀。")
        self.path.parent.mkdir(parents=True, exist_ok=True)

    def _load_or_create(self):
        if self.path.exists():
            try:
                workbook = load_workbook(self.path)
            except Exception as exc:
                raise ExcelStoreError(f"无法打开Excel文件：{exc}") from exc
            if SHEET_NAME not in workbook.sheetnames:
                sheet = workbook.create_sheet(SHEET_NAME)
                self._initialize_sheet(sheet)
            else:
                sheet = workbook[SHEET_NAME]
                existing_headers = tuple(sheet.cell(1, i + 1).value for i in range(len(HEADERS)))
                if sheet.max_row > 1 and existing_headers != HEADERS:
                    raise ExcelStoreError("所选文件中的结果工作表格式不兼容，请选择新的Excel文件。")
                if sheet.max_row == 1 and all(value is None for value in existing_headers):
                    self._initialize_sheet(sheet)
            return workbook, sheet

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = SHEET_NAME
        self._initialize_sheet(sheet)
        return workbook, sheet

    @staticmethod
    def _initialize_sheet(sheet) -> None:
        sheet.append(HEADERS)
        fill = PatternFill("solid", fgColor="1F4E78")
        for cell in sheet[1]:
            cell.fill = fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = "A1:F1"
        widths = (32, 21, 14, 18, 18, 28)
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = width
        sheet.row_dimensions[1].height = 24
        sheet.sheet_view.showGridLines = False

    def append_if_new(
        self,
        domain: str,
        checked_at: str,
        pattern: str,
        prefix: str,
        suffix: str,
        site: str,
    ) -> bool:
        workbook, sheet = self._load_or_create()
        normalized = domain.lower()
        existing = {
            str(sheet.cell(row, 1).value).strip().lower()
            for row in range(2, sheet.max_row + 1)
            if sheet.cell(row, 1).value
        }
        if normalized in existing:
            return False
        try:
            time_value = datetime.fromisoformat(checked_at)
            if time_value.tzinfo is not None:
                time_value = time_value.astimezone().replace(tzinfo=None)
        except ValueError:
            time_value = checked_at
        sheet.append((normalized, time_value, pattern, prefix, suffix, site))
        row = sheet.max_row
        sheet.cell(row, 1).hyperlink = f"https://{normalized}"
        sheet.cell(row, 1).style = "Hyperlink"
        sheet.cell(row, 2).number_format = "yyyy-mm-dd hh:mm:ss"
        sheet.auto_filter.ref = f"A1:F{row}"
        self._atomic_save(workbook)
        return True

    def sync_found_rows(self, rows: list[tuple], site: str) -> int:
        added = 0
        for row in rows:
            domain, checked_at, pattern, prefix, suffix = row[:5]
            saved_site = row[5] if len(row) > 5 and row[5] else site
            if self.append_if_new(domain, checked_at, pattern, prefix, suffix, saved_site):
                added += 1
        return added

    def _atomic_save(self, workbook) -> None:
        fd, temp_name = tempfile.mkstemp(prefix="domain-results-", suffix=".xlsx", dir=self.path.parent)
        os.close(fd)
        try:
            workbook.save(temp_name)
            os.replace(temp_name, self.path)
        except Exception as exc:
            raise ExcelStoreError(f"保存Excel失败：{exc}") from exc
        finally:
            if os.path.exists(temp_name):
                os.unlink(temp_name)


class HistoryExcelStore:
    """把 SQLite 中的全部查询历史导出为可筛选的 Excel 快照。"""

    def __init__(self, path: Path) -> None:
        self.path = Path(path).expanduser()
        if self.path.suffix.lower() != ".xlsx":
            raise ExcelStoreError("已查询记录文件必须使用 .xlsx 后缀。")
        self.path.parent.mkdir(parents=True, exist_ok=True)

    def export(self, rows: list[tuple]) -> int:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = HISTORY_SHEET_NAME
        sheet.append(HISTORY_HEADERS)
        header_fill = PatternFill("solid", fgColor="175CD3")
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.freeze_panes = "A2"
        sheet.sheet_view.showGridLines = False
        widths = (31, 25, 21, 24, 17, 17, 28, 42)
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = width
        sheet.row_dimensions[1].height = 25

        for domain, status, checked_at, pattern, prefix, suffix, detail, site in rows:
            try:
                time_value = datetime.fromisoformat(checked_at)
                if time_value.tzinfo is not None:
                    time_value = time_value.astimezone().replace(tzinfo=None)
            except (TypeError, ValueError):
                time_value = checked_at
            sheet.append(
                (
                    domain,
                    STATUS_LABELS.get(status, status),
                    time_value,
                    pattern,
                    prefix,
                    suffix,
                    site,
                    detail,
                )
            )
            row_index = sheet.max_row
            sheet.cell(row_index, 1).hyperlink = f"https://{domain}"
            sheet.cell(row_index, 1).style = "Hyperlink"
            sheet.cell(row_index, 3).number_format = "yyyy-mm-dd hh:mm:ss"
            for cell in sheet[row_index]:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        sheet.auto_filter.ref = f"A1:H{max(sheet.max_row, 1)}"
        self._atomic_save(workbook)
        return len(rows)

    def _atomic_save(self, workbook) -> None:
        fd, temp_name = tempfile.mkstemp(prefix="domain-history-", suffix=".xlsx", dir=self.path.parent)
        os.close(fd)
        try:
            workbook.save(temp_name)
            os.replace(temp_name, self.path)
        except Exception as exc:
            raise ExcelStoreError(f"保存已查询记录Excel失败：{exc}") from exc
        finally:
            if os.path.exists(temp_name):
                os.unlink(temp_name)
